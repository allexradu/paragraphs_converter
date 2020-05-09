import openpyxl
import pandas as pd
import platform
from openpyxl.styles import Alignment

descriptions = []

table_location = 'excel\\b.xlsx' if platform.system() == 'Windows' else 'excel/b.xlsx'


def read_excel_first_column():
    # Reading first column of a local excel file
    try:
        if platform.system() == 'Windows':
            df = pd.read_excel(table_location, sheet_name = 0)
        else:
            df = pd.read_excel(table_location, sheet_name = 0)

        print('Excel Read Complete!')

        product_names = df['Description'].tolist()

        return product_names
    except:
        print('Excel File NOT READ. Name your file "a.xls" with the first column "Name"')
        print('and place it in the same directory and the bot file.')


def add_data_to_workbook():
    # loading the workbook
    wb = openpyxl.load_workbook(table_location)

    # wb.active returns a Worksheet object
    ws = wb.active

    row_index = 2
    for each_row in descriptions:
        # starting on cell B2 in the workbook
        key = 'B{index}'.format(index = row_index)
        # going down in the B column
        ws[key] = each_row
        # wrapping multiple lines into one cell
        ws[key].alignment = Alignment(wrapText = True)
        # incrementing index
        row_index += 1
    # saving workbook
    wb.save(table_location)
